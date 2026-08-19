"""
build_directory.py
-------------------
Turns raw scraped data (SoundCloud track search, Fiverr gig search, Google
Places / Maps crawler -- all in the shape produced by common Apify actors)
into a single curated directory.json for the HipHopLord site.

FOLDER LAYOUT EXPECTED (same names Apify/your scrapes already used):
    Producers/            -> raw SoundCloud track-search JSON files
    Sound Engineers/       -> raw Fiverr gig-search JSON files
    Visuals & Design/      -> raw Fiverr gig-search JSON files
    studios_usa/           -> raw Google Places crawler JSON files

Each folder may contain any number of .json files, each holding a JSON
array of raw items straight out of the scraper -- nothing needs to be
cleaned by hand first.

WHY THIS VERSION IS DIFFERENT FROM A SIMPLE "COPY EVERYTHING OVER" SCRIPT
  - SoundCloud search results are individual TRACKS, not producer profiles,
    and a lot of them are not actually beats for sale (label reissue
    accounts, an artist's own released songs, duplicate uploads). This
    script groups tracks by uploader and only keeps uploaders whose
    catalog actually looks like beat/instrumental content, then shows
    one representative track per producer instead of a spammy track list.
  - Fiverr search results mix in loosely related gigs (a "cinematic game
    music" gig showing up under a hip hop search, for example). This
    script filters by genre tag + title keywords before anything is
    published.
  - Google Places results include a few non-studio categories (record
    stores, business centers) that slipped into the "recording studio"
    search. Those are dropped.
  - Every category is deduplicated (by seller ID, place ID, or track URL)
    so re-running the same search twice doesn't create repeat cards.

Run it with:
    python build_directory.py
It writes directory.json next to this script.
"""

import json
import os
import re
import urllib.parse
from collections import defaultdict
from datetime import datetime

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------

FIVERR_AFFILIATE_ID = "45990"  # replace with your real Fiverr affiliate/bta ID

FOLDERS = {
    "producers": ["Producers"],
    "engineers": ["Sound Engineers"],
    "visuals": ["Visuals & Design", "(Visuals & Design"],  # tolerate the stray "(" typo
    "studios": ["studios_usa"],
}

HIP_HOP_GENRES = {"hip_hop", "rap", "trap", "drill", "lofi"}
GENRE_PRIORITY = ["trap", "drill", "rap", "hip_hop", "r_b", "lofi", "edm", "pop", "rock"]
ENGINEER_KEYWORDS = re.compile(r"\bmix|\bmaster|\btune|\bvocal", re.I)
BEAT_SIGNAL = re.compile(
    r"type beat|instrumental|prod\.?\s*by|beatmaker|free beat|beat ?tape|riddim|"
    r"boom ?bap|\bbeat\b",
    re.I,
)
MAX_REASONABLE_FIVERR_PRICE = 400  # filters out obvious outlier/mispriced gigs

# SoundCloud uploaders that pass the automated BEAT_SIGNAL check (title/tag/
# description keywords, or a "beatmaker"-style username) but turned out on
# manual review to NOT actually be producers selling beats -- false positives
# the regex can't catch on its own. Reviewed by hand against their full track
# list before exclusion; reasons kept here for future re-review:
#   - "Emad Saad Hip Hop": rapper/emcee releasing his own vocal songs
#     ("Heavyweight Lyricist", features, verses) -- matched only because
#     "boom bap" is used as his genre/style tag, not because he sells beats.
#   - "Dgunzbeatz": mixing & mastering engineer ("M&M BY DGUNZ" credited on
#     other artists' finished vocal songs) -- matched via a couple of
#     "instrumental"-tagged tracks and one "prod by" credit, but the bulk of
#     the catalog is mix/master work, not original instrumentals for sale.
#   - "Lance O Smith": every track carries the same copy-pasted "Instrumental
#     hiphop..." boilerplate description (a SoundCloud discovery/SEO tactic),
#     but the actual titles ("Bad Bunny", "My Boo", "Track 1"..."Track 9")
#     read as a DJ/remix repost account, not original productions.
#   - "MagroTheHipHopArtist": rapper/emcee (rap-battle cypher entries,
#     features, a Mac Miller tribute); one track's title credits a *different*
#     person as producer ("Produced by @Dangertheproducer"), which is what
#     tripped the filter.
EXCLUDED_PRODUCER_USERNAMES = {
    "Emad Saad Hip Hop",
    "Dgunzbeatz",
    "Lance O Smith",
    "MagroTheHipHopArtist",
}

# Minimal ISO 3166-1 alpha-2 -> country name map, covering the codes that show
# up in Fiverr/SoundCloud data. Falls back to the raw code for anything else.
COUNTRY_NAMES = {
    "US": "United States", "GB": "United Kingdom", "CA": "Canada", "AU": "Australia",
    "DE": "Germany", "FR": "France", "IT": "Italy", "ES": "Spain", "PT": "Portugal",
    "NL": "Netherlands", "BE": "Belgium", "PL": "Poland", "RO": "Romania",
    "BG": "Bulgaria", "GR": "Greece", "RS": "Serbia", "SI": "Slovenia",
    "CZ": "Czechia", "LT": "Lithuania", "EE": "Estonia", "UA": "Ukraine",
    "LK": "Sri Lanka", "PK": "Pakistan", "IN": "India", "BD": "Bangladesh",
    "PH": "Philippines", "ID": "Indonesia", "TH": "Thailand", "NG": "Nigeria",
    "KE": "Kenya", "ZA": "South Africa", "MA": "Morocco", "ET": "Ethiopia",
    "AR": "Argentina", "BR": "Brazil", "CL": "Chile", "BO": "Bolivia",
    "PE": "Peru", "CO": "Colombia", "MX": "Mexico", "DO": "Dominican Republic",
    "AZ": "Azerbaijan", "TJ": "Tajikistan",
}


def country_name(code):
    if not code:
        return ""
    return COUNTRY_NAMES.get(code.strip().upper(), code)


def parse_soundcloud_tags(tag_list):
    """SoundCloud tag_list mixes quoted multi-word tags with bare single
    words, e.g.  '"Mobb Deep" "Hip Hop" Remix Rap'. A plain .split() breaks
    the quoted phrases apart -- this keeps them intact."""
    if not tag_list:
        return []
    tokens = re.findall(r'"([^"]+)"|(\S+)', tag_list)
    tags = [a or b for a, b in tokens]
    return [t.strip().lower() for t in tags if t.strip()]


def sorted_by_genre_priority(genres):
    genres = list(genres)
    genres.sort(key=lambda g: GENRE_PRIORITY.index(g) if g in GENRE_PRIORITY else 99)
    return [g.replace("_", " ") for g in genres]


# ---------------------------------------------------------------------------
# GENERIC HELPERS
# ---------------------------------------------------------------------------

def load_json_files(folder_names, base_dir):
    items = []
    for name in folder_names:
        path = os.path.join(base_dir, name)
        if not os.path.isdir(path):
            continue
        for filename in sorted(os.listdir(path)):
            if not filename.endswith(".json"):
                continue
            try:
                with open(os.path.join(path, filename), "r", encoding="utf-8") as f:
                    data = json.load(f)
            except Exception as error:
                print(f"  (skipped {filename}: {error})")
                continue
            if isinstance(data, list):
                items.extend(data)
    return items


def make_fiverr_affiliate_link(raw_url):
    """Build a Fiverr affiliate deep link to a specific gig page.

    IMPORTANT: Fiverr's go.fiverr.com/visit/ redirector only honors the
    param name "landingPage" (camelCase), NOT "landing_page" (snake_case) --
    and it expects the target URL to be percent-encoded TWICE. Confirmed by
    comparing against a real link generated from the Fiverr affiliate
    dashboard's own "LP URL" tool:

        ...&landingPage=https%253A%252F%252Fwww.fiverr.com%252F<user>%252F<gig>

    Using "landing_page" (single-encoded, as this function did before) is
    silently ignored by Fiverr and falls back to the generic homepage --
    that was the root cause of every Engineers/Visuals card resolving to
    the same non-specific Fiverr URL.
    """
    if not raw_url or "fiverr.com" not in raw_url:
        return raw_url or ""
    clean_url = raw_url.split("?")[0]
    encoded_once = urllib.parse.quote(clean_url, safe="")
    encoded_twice = urllib.parse.quote(encoded_once, safe="")
    return (
        "https://go.fiverr.com/visit/"
        f"?bta={FIVERR_AFFILIATE_ID}&brand=fiverrmarketplace&landingPage={encoded_twice}"
    )


def make_soundcloud_embed(track_url):
    if not track_url:
        return ""
    encoded = urllib.parse.quote(track_url, safe="")
    return (
        "https://w.soundcloud.com/player/?url=" + encoded +
        "&color=%23e2a83f&auto_play=false&hide_related=true&show_comments=false"
        "&show_user=true&show_reposts=false&visual=false"
    )


def title_case_words(value):
    return value.strip() if value else ""


def clean_producer_name(username):
    """SoundCloud usernames are often 'Name | extra tagline' or
    'Name (extra description)' -- keep just the clean name for display."""
    if not username:
        return "Unknown Artist"
    name = re.split(r"\s*[|(]", username)[0].strip()
    return name or username.strip()


def format_count(n):
    """1234 -> '1.2K', 4500000 -> '4.5M', small numbers unchanged."""
    if n is None:
        return None
    n = int(n)
    if n >= 1_000_000:
        return f"{n / 1_000_000:.1f}M".replace(".0M", "M")
    if n >= 1_000:
        return f"{n / 1_000:.1f}K".replace(".0K", "K")
    return str(n)


def format_upload_date(iso_str):
    """'2026-08-05T03:31:32Z' -> 'Aug 2026' for a lightweight recency signal."""
    if not iso_str:
        return None
    try:
        dt = datetime.strptime(iso_str[:10], "%Y-%m-%d")
        return dt.strftime("%b %Y")
    except (ValueError, TypeError):
        return None


LANGUAGE_NAMES = {
    "en": "English", "es": "Spanish", "fr": "French", "de": "German",
    "it": "Italian", "pt": "Portuguese", "nl": "Dutch", "ru": "Russian",
    "ar": "Arabic", "tr": "Turkish", "pl": "Polish", "ro": "Romanian",
    "id": "Indonesian", "hi": "Hindi", "zh-cn": "Chinese", "zh": "Chinese",
    "ja": "Japanese", "ko": "Korean", "sv": "Swedish", "el": "Greek",
    "uk": "Ukrainian", "he": "Hebrew", "vi": "Vietnamese", "th": "Thai",
    "fa": "Persian", "cs": "Czech", "fil": "Filipino", "af": "Afrikaans",
    "bg": "Bulgarian", "bn": "Bengali", "ca": "Catalan", "si": "Sinhala",
    "sr": "Serbian", "ta": "Tamil", "ur": "Urdu",
}


def format_languages(codes):
    if not codes:
        return None
    seen = []
    for c in codes:
        name = LANGUAGE_NAMES.get(c.lower(), c.upper())
        if name not in seen:
            seen.append(name)
        if len(seen) == 3:
            break
    return ", ".join(seen)


def summarize_opening_hours(opening_hours):
    """List of {"day": "Monday", "hours": "Open 24 hours"} -> a short,
    always-accurate (non-time-sensitive) one-line summary. We deliberately
    do not compute "open now" since that would need a live clock and this
    is a static build."""
    if not opening_hours:
        return None
    by_day = {h.get("day"): h.get("hours") for h in opening_hours if h.get("day")}
    days_order = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"]
    values = [by_day.get(d) for d in days_order]
    if not any(values):
        return None
    if all(v == values[0] for v in values if v):
        return f"{values[0]}, every day" if values[0] else None
    weekdays = values[:5]
    weekend = values[5:]
    if all(v == weekdays[0] for v in weekdays if v) and all(v == weekend[0] for v in weekend if v) and weekdays[0] and weekend[0]:
        return f"Mon–Fri {weekdays[0]} · Sat–Sun {weekend[0]}"
    return "Hours vary by day — see Google Maps listing"


def extract_amenity_badges(additional_info):
    """Google Places 'additionalInfo' is a dict of category -> list of
    {label: bool}. Surface only a small, useful set of true amenities."""
    if not additional_info:
        return []
    wanted = {
        "Wi-Fi": "Wi-Fi",
        "Free parking lot": "Free Parking",
        "On-site parking": "On-site Parking",
        "Valet parking": "Valet Parking",
        "Wheelchair accessible entrance": "Wheelchair Accessible",
    }
    found = []
    for _group, items in additional_info.items():
        for item in items or []:
            for label, value in item.items():
                if value and label in wanted and wanted[label] not in found:
                    found.append(wanted[label])
    return found[:3]


# ---------------------------------------------------------------------------
# PRODUCERS  (SoundCloud track search -> one card per producer)
# ---------------------------------------------------------------------------

def build_producers(base_dir):
    raw_tracks = load_json_files(FOLDERS["producers"], base_dir)
    by_user = defaultdict(list)
    for track in raw_tracks:
        user = track.get("user") or {}
        username = user.get("username") or "Unknown"
        by_user[username].append(track)

    entries = []
    for username, tracks in by_user.items():
        if username in EXCLUDED_PRODUCER_USERNAMES:
            continue

        def has_signal(t):
            haystack = " ".join([
                t.get("title") or "", t.get("tag_list") or "", t.get("description") or "",
            ])
            return bool(BEAT_SIGNAL.search(haystack))

        signal_tracks = [t for t in tracks if has_signal(t)]
        signal_ratio = len(signal_tracks) / len(tracks) if tracks else 0
        username_signals = bool(BEAT_SIGNAL.search(username))

        # Keep this uploader only if their catalog reads like actual beat/
        # instrumental content -- either most of their tracks show beat
        # signals, or the account name itself makes it obvious (e.g. a
        # username containing "Beatmaker" or "Type Beat").
        if signal_ratio < 0.3 and not username_signals:
            continue

        pool = signal_tracks or tracks
        best = max(pool, key=lambda t: t.get("likes_count") or 0)

        genre = best.get("genre") or "Hip Hop"
        clean_name = clean_producer_name(username)
        tags = parse_soundcloud_tags(best.get("tag_list"))[:4] or [genre.lower()]

        user = best.get("user") or {}
        location = ", ".join(filter(None, [user.get("city"), country_name(user.get("country_code"))]))

        extra_stats = []
        followers = format_count(user.get("followers_count"))
        if followers:
            extra_stats.append(f"{followers} followers")
        plays = format_count(best.get("playback_count"))
        if plays:
            extra_stats.append(f"{plays} plays")
        uploaded = format_upload_date(best.get("created_at"))
        if uploaded:
            extra_stats.append(f"uploaded {uploaded}")

        badges = ["✓ Verified SoundCloud Account"] if user.get("verified") else []

        entries.append({
            "id": f"prod-{len(entries) + 1:03d}",
            "category": "producers",
            "categoryLabel": "Producers & Beatmakers",
            "title": title_case_words(best.get("title")) or "Hip Hop Instrumental",
            "provider_name": clean_name,
            "location": location,
            "price": "Contact for licensing",
            "description": (best.get("description") or "").strip()[:220] or
                            f"Original {genre.lower()} production from {clean_name}, sampled from their SoundCloud catalog.",
            "tags": tags,
            "link": best.get("permalink_url", ""),
            "linkLabel": "Listen on SoundCloud",
            "embed_url": make_soundcloud_embed(best.get("permalink_url", "")),
            "image": best.get("artwork_url") or "",
            "likes": best.get("likes_count") or 0,
            "extraStats": extra_stats,
            "badges": badges,
            "featured": (best.get("likes_count") or 0) > 50,
            "sample": False,
        })

    return entries


# ---------------------------------------------------------------------------
# MANUALLY-SOURCED PRODUCERS  (found via Google search discovery, e.g.
# `site:soundcloud.com "type beat" "prod by" hip hop`, then verified by hand
# by reading each account's bio and at least one track's real play/like
# counts and artwork on soundcloud.com -- not run through the automated
# BEAT_SIGNAL filter since there's no bulk raw-track JSON file for these.
# Each entry cites the exact track used and the stats at verification time
# so this batch can be re-checked or re-scraped properly later.
# ---------------------------------------------------------------------------

MANUAL_PRODUCER_ADDITIONS = [
    {
        # Verified: bio explicitly "Beatmaker Belge" (Belgium), beatstore in
        # bio, 51-track catalog of consistent "X Type Beat" titles.
        "title": 'Rick Ross Type Beat - "Luxurious" [2024] | Rich Trap Type Beat',
        "provider_name": "Sly The Beatmaker",
        "location": "Belgium",
        "tags": ["trap", "type beat"],
        "link": "https://soundcloud.com/sly-the-beat-maker/rick-ross-type-beat-luxurious-2024-rich-trap-type-beat",
        "image": "https://i1.sndcdn.com/artworks-zKMXCD8Fv5ZXWpIa-geKQbQ-t500x500.jpg",
        "likes": 4,
        "genre": "Trap",
    },
    {
        # Verified: bio "African beatmaker. Trap beats. Type beats...",
        # 25-track catalog of "X Type Beat (Prod. UnplugBeats)" titles.
        "title": "[FREE] Gunna x Young Thug DOLLAZ ON MY HEAD Type Beat",
        "provider_name": "UnplugBeats",
        "location": "Nigeria",
        "tags": ["trap", "type beat", "free beat"],
        "link": "https://soundcloud.com/beatsunplug/free-gunna-x-young-thug-dollaz",
        "image": "https://i1.sndcdn.com/artworks-000801320986-xs45vu-t500x500.jpg",
        "likes": 7,
        "genre": "Hip Hop",
    },
    {
        # Verified: bio "Perfil do BeatMaker 'Mayck Beats' (BRAZIL)",
        # 53-track catalog, strong engagement (6,125 plays on this track).
        "title": "Trap Funk Type Beat Orochi x Travis Scott (Instrumental Hip Hop and Trap 808)",
        "provider_name": "Mayck Beats",
        "location": "Bahia, Brazil",
        "tags": ["trap type beat"],
        "link": "https://soundcloud.com/mayckmc/trap-funk-type-beat-orochi-x-travis-scott-instrumental-hip-hop-and-trap-808-mayck-beats",
        "image": "https://i1.sndcdn.com/artworks-HcwDA773i1U86Zqz-RBcwgg-t500x500.jpg",
        "likes": 113,
        "genre": "Trap",
        "extra_stats": ["6.1K plays"],
    },
    {
        # Verified: bio "I am a producer #hubertbeat #beats #instrumentals
        # #freebeats". Thin catalog (3 tracks) but genuine, real engagement.
        "title": "Sarkodie instrumental-Type Beat (Prod By HubertBeat)",
        "provider_name": "HubertBeat",
        "location": "Accra, Ghana",
        "tags": ["beats", "afrobeat", "instrumental"],
        "link": "https://soundcloud.com/hubertbeat/sarkodie-instrumental-type-beat-prod-by-hubertbeat",
        "image": "https://i1.sndcdn.com/artworks-000637963720-8j67hj-t500x500.jpg",
        "likes": 8,
        "genre": "Hip Hop",
    },
    {
        # Verified: track description explicitly asks for "YcBeatmaker"
        # credit, lists Instagram/email/YouTube contact -- real beatmaker.
        "title": "Instru Sad Hard Rap | Trap Type Beat Instrumental 2020 [Prod By YCBeatMaker]",
        "provider_name": "YC BeatMaker",
        "location": "",
        "tags": ["trap", "instrumental", "hip hop"],
        "link": "https://soundcloud.com/yassine-chohra/projet-16-by-ycbeatmaker-mp3",
        "image": "https://i1.sndcdn.com/artworks-bwfPqQAylu6AQNid-BL0kqA-t500x500.jpg",
        "likes": 5,
        "genre": "Hip Hop",
    },
    {
        # Verified: profile "APhoniC Beatz", Rio de Janeiro, 1,222 followers,
        # 3,324 plays on this track.
        "title": "YG Type Beat #1",
        "provider_name": "APhoniC Beatz",
        "location": "Rio de Janeiro, Brazil",
        "tags": ["trap", "rap", "hip hop"],
        "link": "https://soundcloud.com/aphonicbeats/yg-type-beat-1",
        "image": "https://i1.sndcdn.com/artworks-000021780032-c4w9yz-t500x500.jpg",
        "likes": 28,
        "genre": "Hip Hop",
        "extra_stats": ["1.2K followers", "3.3K plays"],
    },
    {
        # Verified: description promotes "purchasing instrumentals",
        # 5,499 plays on this track -- established commercial beatmaker.
        "title": "Instrumental Rap Hip Hop Boom Bap Old School Gratis Uso Libre Free Beat",
        "provider_name": "BeatMaker Beatz",
        "location": "",
        "tags": ["boom bap", "old school", "instrumental"],
        "link": "https://soundcloud.com/beatmaker_beatz/instrumental-rap-hip-hop-boom-bap-old-school-gratis-uso-libre-free-beat",
        "image": "https://i1.sndcdn.com/artworks-000198821432-vpb6wl-t500x500.jpg",
        "likes": 76,
        "genre": "Hip Hop",
        "extra_stats": ["5.5K plays"],
    },
    {
        # Verified: extensive on-topic tag list (boom bap, old school
        # instrumental, rap instrumental...), consistent beat-catalog brand.
        "title": "Old School Boom Bap Beat | Hip Hop Rap Instrumental - Lifestyle",
        "provider_name": "Smallz Productions",
        "location": "",
        "tags": ["boom bap", "old school instrumental", "rap beat"],
        "link": "https://soundcloud.com/user-972973697/old-school-boom-bap-beat-hip-hop-rap-instrumental-lifestyle",
        "image": "https://i1.sndcdn.com/artworks-000478951482-62vbfx-t500x500.jpg",
        "likes": 13,
        "genre": "Hip Hop",
    },
]


def build_manual_producer_additions(start_index):
    entries = []
    for i, item in enumerate(MANUAL_PRODUCER_ADDITIONS):
        entries.append({
            "id": f"prod-{start_index + i:03d}",
            "category": "producers",
            "categoryLabel": "Producers & Beatmakers",
            "title": title_case_words(item["title"]),
            "provider_name": item["provider_name"],
            "location": item["location"],
            "price": "Contact for licensing",
            "description": f"Original {item['genre'].lower()} production from {item['provider_name']}, "
                            f"sampled from their SoundCloud catalog.",
            "tags": item["tags"][:4],
            "link": item["link"],
            "linkLabel": "Listen on SoundCloud",
            "embed_url": make_soundcloud_embed(item["link"]),
            "image": item["image"],
            "likes": item["likes"],
            "extraStats": item.get("extra_stats", []),
            "badges": [],
            "featured": item["likes"] > 50,
            "sample": False,
        })
    return entries


def generate_fiverr_curator_note(entry):
    """Rule-based, honest 'why this listing is here' note for a Fiverr-sourced
    engineer/designer entry. Built entirely from real fields already on the
    entry (rating, reviewsCount, extraStats, badges) -- no invented claims.
    Used as the base layer for all engineer/designer listings; higher-volume
    featured listings additionally get a hand-written note via
    FIVERR_FEATURED_CURATOR_NOTES that overrides this."""
    rating = entry.get("rating")
    reviews = entry.get("reviewsCount") or 0
    extra_stats = entry.get("extraStats") or []
    badges = entry.get("badges") or []
    location = entry.get("location")
    noun = "engineer" if entry.get("category") == "engineers" else "designer"
    based_in = f" based in {location}" if location else ""

    delivery = next((s for s in extra_stats if "delivery" in s), None)
    languages = next((s for s in extra_stats if s.startswith("speaks")), None)
    is_pro = "PRO Verified Seller" in badges

    if rating and reviews >= 200:
        base = f"A well-reviewed Fiverr {noun}{based_in}: {rating}★ across {reviews} reviews"
    elif rating and reviews >= 30:
        base = f"A Fiverr {noun}{based_in} with a solid review history: {rating}★ from {reviews} reviews"
    elif rating and reviews > 0:
        base = (f"A Fiverr {noun}{based_in} with {reviews} review"
                f"{'s' if reviews != 1 else ''} at {rating}★ so far")
    else:
        base = (f"A Fiverr {noun}{based_in} with limited review history on this specific gig "
                f"-- worth checking their full profile before booking")

    extras = []
    if is_pro:
        extras.append("Fiverr Pro-verified")
    if delivery:
        extras.append(delivery)
    if languages:
        extras.append(languages)

    return base + (f" ({', '.join(extras)})." if extras else ".")


# Hand-written, deeper editorial notes for the 25 "featured" engineer/designer
# listings (rating >= 4.9 and reviews >= 500 -- see build_fiverr_category()).
# Every sentence below is grounded strictly in the listing's own real Fiverr
# data (rating, review count, delivery time, languages, price, PRO status,
# style tag). Applied as an override after generate_fiverr_curator_note() has
# already populated every other engineer/designer listing, so no entry is
# left without a note -- the featured ones simply get more editorial depth.
FIVERR_FEATURED_CURATOR_NOTES = {
    "eng-001": "With just under 7,000 reviews at 4.9★, this is the single "
               "most-reviewed mixing engineer in our directory -- a 2-day "
               "turnaround and pricing starting at $45 make it a low-risk "
               "first booking for artists who want a proven track record.",
    "eng-002": "A Fiverr Pro-verified engineer with a perfect 5★ across "
               "4,117 reviews. Pricing starts higher (from $135), which tracks "
               "with the 'major label standard' positioning and PRO status.",
    "eng-003": "4.9★ across 3,204 reviews at a starting price of just $20 "
               "-- one of the more budget-friendly options among the "
               "higher-volume engineers, specializing specifically in rap "
               "vocal mixing.",
    "eng-005": "A Fiverr Pro-verified engineer fluent in German, English and "
               "Turkish, with 2,405 reviews at 4.9★ -- a solid option for "
               "artists who want multilingual communication during mixing.",
    "eng-007": "A perfect 5★ rating across 1,595 reviews, with pricing "
               "starting at $35 -- strong value for a proven mixing and "
               "mastering track record.",
    "eng-008": "Fiverr Pro-verified with 1,596 reviews at 4.9★. Based in "
               "the US, which can simplify time-zone coordination for artists "
               "working with US-based collaborators.",
    "eng-009": "A 1-day delivery window paired with a perfect 5★ across "
               "1,442 reviews -- one of the fastest turnarounds among our "
               "higher-volume engineers.",
    "eng-010": "Fiverr Pro-verified, 4.9★ across 1,291 reviews, and "
               "trilingual (Greek, English, Spanish) -- a well-rounded, "
               "established profile.",
    "eng-011": "4.9★ across 848 reviews at a starting price of $15 -- "
               "among the most affordable options in this category with a "
               "genuinely large review history behind it.",
    "eng-012": "A perfect 5★ across 793 reviews, 1-day delivery, and a "
               "$15 starting price -- a fast, budget-friendly option with a "
               "strong track record.",
    "vis-001": "The most-reviewed cover art designer in our directory: 5,439 "
               "reviews at 4.9★, specializing in photographic-style album "
               "art with a 2-day turnaround.",
    "vis-002": "4,599 reviews at 4.9★ with a 1-day delivery window and "
               "pricing starting at just $20 -- strong volume and speed at a "
               "low price point, plus a 9-image portfolio to review beforehand.",
    "vis-003": "4,175 reviews at 4.9★ and 1-day delivery -- one of the "
               "fastest-turnaround, highest-volume designers in the "
               "photographic-style category.",
    "vis-004": "4,047 reviews at 4.9★ in a typographic cover-art style. "
               "Note the longer 14-day delivery window compared to most other "
               "designers here -- worth confirming timing directly if it matters.",
    "vis-005": "3,306 reviews at 4.9★, a 9-image portfolio, and a 3-day "
               "turnaround -- a well-established photographic-style cover artist.",
    "vis-006": "2,900 reviews at 4.9★ with pricing starting at $30 -- "
               "strong review volume at a moderate price point.",
    "vis-007": "Fiverr Pro-verified, 4.9★ across 1,467 reviews, "
               "specializing in vintage-collage-style album art -- a distinct "
               "visual niche among our designers.",
    "vis-008": "1,369 reviews at 4.9★ with a 2-day turnaround and a $15 "
               "starting price -- one of the most affordable high-volume "
               "designers in the directory.",
    "vis-009": "1,191 reviews at 4.9★, specializing in collage-style "
               "artwork with a 7-image portfolio to preview before booking.",
    "vis-011": "1,148 reviews at 4.9★, 1-day delivery, and a $10 starting "
               "price -- the lowest entry price among our featured designers, "
               "with a genuine review history behind it.",
    "vis-012": "1,146 reviews at 4.9★ and an 8-image portfolio at a $10 "
               "starting price -- solid value for artists on a tighter budget.",
    "vis-013": "919 reviews at 4.9★ with 2-day delivery and a $10 "
               "starting price.",
    "vis-015": "585 reviews at 4.9★, trilingual (English, Arabic, "
               "French), specializing in typographic album art.",
    "vis-017": "A perfect 5★ across 511 reviews, specializing in "
               "illustrative-style cover art -- a good pick for artists "
               "wanting hand-drawn or painted visuals rather than photo "
               "composites.",
    "vis-018": "516 reviews at 4.9★, photographic-style cover art, "
               "starting at $30.",
}


# ---------------------------------------------------------------------------
# ENGINEERS + VISUALS  (Fiverr gig search)
# ---------------------------------------------------------------------------

def build_fiverr_category(base_dir, folder_key, category, label, id_prefix,
                           require_hip_hop_genre, require_engineer_keywords,
                           genre_field="genre"):
    raw_gigs = load_json_files(FOLDERS[folder_key], base_dir)
    by_seller = {}

    for gig in raw_gigs:
        title = (gig.get("title") or "").strip()
        if not title:
            continue

        price = gig.get("priceFrom")
        if price is not None and price > MAX_REASONABLE_FIVERR_PRICE:
            continue

        genres = set((gig.get("attributes") or {}).get(genre_field) or [])
        if require_hip_hop_genre and not (genres & HIP_HOP_GENRES):
            continue
        if require_engineer_keywords and not ENGINEER_KEYWORDS.search(title):
            continue

        seller_id = gig.get("sellerId") or gig.get("sellerName")
        score = (gig.get("rating") or 0) * (gig.get("reviewsCount") or 0)

        # Keep only the best-scoring gig per seller, so one popular seller
        # with many gigs doesn't crowd out everyone else.
        existing = by_seller.get(seller_id)
        if existing is None or score > existing["_score"]:
            gig["_score"] = score
            by_seller[seller_id] = gig

    ranked = sorted(by_seller.values(), key=lambda g: g["_score"], reverse=True)

    entries = []
    for gig in ranked:
        attributes = gig.get("attributes") or {}
        genres = set(attributes.get(genre_field) or [])
        design_styles = attributes.get("design_style") or []

        # Cover art gigs are more usefully tagged by visual style than by
        # musical genre (every gig in this file targets hip hop already).
        if design_styles:
            tags = [s.replace("_", " ") for s in design_styles][:3]
        else:
            tags = sorted_by_genre_priority(genres & HIP_HOP_GENRES or genres)[:4]
        tags = tags or ["hip hop"]

        price = gig.get("priceFrom")
        rating = gig.get("rating")
        reviews = gig.get("reviewsCount") or 0
        seller_level = {
            "top_rated_seller": "Fiverr Top Rated Seller",
            "level_two_seller": "Fiverr Level 2 Seller",
            "level_one_seller": "Fiverr Level 1 Seller",
        }.get(gig.get("sellerLevel"), "Fiverr Seller")

        description = seller_level
        if reviews:
            description += f" · {reviews} reviews"
        if rating:
            description += f" · {rating}★"

        extra_stats = []
        packages = gig.get("packages") or []
        delivery_days = packages[0].get("durationDays") if packages else None
        if delivery_days:
            extra_stats.append(f"{delivery_days}-day delivery")
        languages = format_languages(gig.get("sellerLanguages"))
        if languages:
            extra_stats.append(f"speaks {languages}")
        gallery_count = len(gig.get("images") or [])
        if gallery_count > 1:
            extra_stats.append(f"{gallery_count} portfolio images")

        badges = []
        if gig.get("isPro"):
            badges.append("PRO Verified Seller")

        entries.append({
            "id": f"{id_prefix}-{len(entries) + 1:03d}",
            "category": category,
            "categoryLabel": label,
            "title": title_case_words(gig.get("title")),
            "provider_name": gig.get("sellerDisplayName") or gig.get("sellerName") or "",
            "location": country_name(gig.get("sellerCountry")),
            "price": f"From ${price:g}" if price is not None else "",
            "description": description,
            "tags": tags,
            "link": make_fiverr_affiliate_link(gig.get("gigUrl", "")),
            "linkLabel": "View Gig on Fiverr",
            "image": (gig.get("images") or [None])[0] or gig.get("sellerImage") or "",
            "rating": rating,
            "reviewsCount": reviews,
            "extraStats": extra_stats,
            "badges": badges,
            "featured": (rating or 0) >= 4.9 and reviews >= 500,
            "sample": False,
        })

    return entries


def generate_studio_curator_note(entry):
    """Rule-based, honest 'why this listing is here' note for a Google-Places
    studio entry. Built entirely from real fields already on the entry
    (rating, reviewsCount, badges) -- no invented claims. Used as the base
    layer for every studio listing; the 66 'featured' studios (rating >= 4.8
    and reviews >= 100 -- see build_studios()) additionally get a hand-written
    note via STUDIO_FEATURED_CURATOR_NOTES that overrides this."""
    rating = entry.get("rating")
    reviews = entry.get("reviewsCount") or 0
    badges = entry.get("badges") or []
    location = entry.get("location")
    in_location = f" in {location}" if location else ""

    if rating and reviews > 0:
        base = f"Rated {rating}★ from {reviews} Google review{'s' if reviews != 1 else ''}{in_location}."
    else:
        base = (f"No Google reviews yet on this listing{in_location} -- worth confirming "
                 "details directly before booking.")

    if badges:
        if len(badges) == 1:
            amenity_text = badges[0]
        elif len(badges) == 2:
            amenity_text = f"{badges[0]} and {badges[1]}"
        else:
            amenity_text = ", ".join(badges[:-1]) + f", and {badges[-1]}"
        base += f" Listed amenities: {amenity_text}."

    return base


# Hand-written, deeper editorial notes for the 66 "featured" studio listings
# (rating >= 4.8 and reviews >= 100 -- see build_studios()). Every sentence
# is grounded strictly in the listing's own real Google Places data (rating,
# review count, opening hours, photo count, amenity badges, and Google's own
# secondary category tags such as "event venue" or "music producer" on that
# listing). Applied as an override after generate_studio_curator_note() has
# already populated every other studio listing, so no entry is left without
# a note -- the featured ones simply get more editorial depth.
STUDIO_FEATURED_CURATOR_NOTES = {
    "std-005": "A 5★-rated Atlanta studio with 116 Google reviews, open "
               "daily from 9 AM to 11:30 PM and Wi-Fi on site.",
    "std-011": "With 317 Google reviews at 4.8★ and 399 photos documenting "
               "the space, this is one of the more established Atlanta "
               "studios in our directory -- open 24 hours, wheelchair "
               "accessible, with free parking, and it also doubles as an "
               "event venue per its own Google listing.",
    "std-018": "4.9★ from 130 Google reviews in Austin, open daily from 9 "
               "AM to midnight, wheelchair accessible with Wi-Fi on site.",
    "std-024": "A perfect 5★ across 452 Google reviews -- one of the "
               "highest review counts among Baltimore studios in our "
               "directory -- open 24 hours with Wi-Fi on site.",
    "std-027": "4.8★ from 153 Google reviews, open 24 hours daily. Also "
               "operates as a rehearsal space and instrument repair shop, "
               "per its own Google listing.",
    "std-031": "4.9★ from 167 Google reviews. Notably, this is registered "
               "as a non-profit organization on Google -- a different model "
               "from most for-profit studios in this directory -- with "
               "wheelchair access, Wi-Fi and free parking.",
    "std-032": "4.8★ from 184 Google reviews, open 24 hours, wheelchair "
               "accessible with Wi-Fi on site.",
    "std-039": "A perfect 5★ across 121 reviews in Charlotte, open late "
               "(10 AM to 4 AM daily). Its Google listing also covers "
               "acoustical consulting and music education, alongside "
               "production and recording.",
    "std-041": "5★ from 118 Google reviews, open 24 hours, wheelchair "
               "accessible with Wi-Fi on site.",
    "std-043": "4.9★ from 147 Google reviews, open daily until 2 AM, with "
               "wheelchair access, Wi-Fi and free parking.",
    "std-044": "4.8★ from 189 Google reviews. Its Google listing also "
               "covers music education (learning center) alongside "
               "recording and production, with both free and on-site parking.",
    "std-045": "4.8★ from 127 Google reviews, open 24 hours, wheelchair "
               "accessible with Wi-Fi on site.",
    "std-052": "4.9★ from 123 Google reviews in Chicago, open 24 hours. "
               "Its listing spans audio-visual consulting and video "
               "duplication alongside recording, wheelchair accessible with "
               "Wi-Fi on site.",
    "std-062": "A perfect 5★ across 123 reviews, with 216 photos "
               "documenting the space on Google -- more than most studios in "
               "this directory. Open daily 7 AM to 10 PM, wheelchair "
               "accessible with Wi-Fi.",
    "std-068": "4.9★ from 131 Google reviews in Dallas, open daily 10 AM "
               "to 7 PM, wheelchair accessible with Wi-Fi on site.",
    "std-070": "4.8★ from 178 Google reviews, open daily until midnight, "
               "with 166 photos on its Google listing and wheelchair access.",
    "std-071": "4.8★ from 100 Google reviews, open daily 8 AM to 11 PM. "
               "Also lists music management and promotion services alongside "
               "recording.",
    "std-076": "A perfect 5★ across 108 Google reviews in Denver, open 24 "
               "hours daily, wheelchair accessible.",
    "std-082": "4.8★ from 141 Google reviews, open 24 hours daily.",
    "std-083": "4.8★ from 113 Google reviews, open 24 hours daily.",
    "std-085": "4.8★ from 311 Google reviews -- one of the higher review "
               "counts in the Denver area -- with 226 photos on its listing. "
               "Also operates as a music conservatory and instrument store "
               "alongside rehearsal space, wheelchair accessible.",
    "std-089": "4.9★ from 117 Google reviews in Detroit, with Wi-Fi on site.",
    "std-090": "4.8★ from 117 Google reviews, open daily noon to midnight, "
               "with Wi-Fi on site.",
    "std-098": "A perfect 5★ across 338 Google reviews -- one of the "
               "strongest review counts in our directory -- open 24 hours, "
               "wheelchair accessible with Wi-Fi on site.",
    "std-100": "A perfect 5★ across 132 Google reviews, with defined "
               "weekday (11 AM–10 PM) and weekend (8 AM–5 PM) hours -- useful "
               "for artists who want to plan a session in advance.",
    "std-101": "4.8★ from 111 Google reviews with 313 photos documenting "
               "the space -- among the most photo-documented studios in this "
               "directory. Open 24 hours, wheelchair accessible with Wi-Fi.",
    "std-108": "4.9★ from 122 Google reviews in Houston, wheelchair "
               "accessible with Wi-Fi on site.",
    "std-109": "4.9★ from 136 Google reviews, with wheelchair access, "
               "Wi-Fi and free parking.",
    "std-118": "A perfect 5★ across 172 Google reviews in Las Vegas, open "
               "24 hours, wheelchair accessible with Wi-Fi on site.",
    "std-119": "A perfect 5★ across 122 Google reviews, open 24 hours. "
               "Its listing also covers video production alongside "
               "recording, wheelchair accessible with Wi-Fi.",
    "std-122": "A perfect 5★ across 104 Google reviews, open 24 hours, "
               "with wheelchair access, Wi-Fi and free parking.",
    "std-123": "A perfect 5★ across 140 Google reviews, with defined "
               "weekday hours (10 AM–8 PM, closed weekends) and free parking "
               "on site.",
    "std-127": "With 844 Google reviews at 4.9★, this is one of the "
               "most-reviewed studios in the entire directory -- open 24 "
               "hours, wheelchair accessible with Wi-Fi, and extensively "
               "documented with 315 photos.",
    "std-135": "A perfect 5★ across 113 Google reviews in Los Angeles, "
               "open daily 8 AM to 1 AM, wheelchair accessible with Wi-Fi.",
    "std-136": "A perfect 5★ across 117 Google reviews, open 24 hours. "
               "Its listing spans acoustical consulting and music management "
               "alongside recording and production.",
    "std-137": "409 Google reviews at 4.9★ and 475 photos -- one of the "
               "most extensively documented studios in our directory. Open "
               "24 hours, wheelchair accessible, with Wi-Fi and free parking.",
    "std-138": "240 Google reviews at 4.9★, open 24 hours with wheelchair "
               "access, Wi-Fi and free parking. Its listing also covers "
               "video production alongside recording.",
    "std-140": "231 Google reviews at 4.8★, with 393 photos on its "
               "listing. Open 24 hours, wheelchair accessible, with Wi-Fi "
               "and free parking.",
    "std-143": "4.8★ from 119 Google reviews, open 24 hours, wheelchair "
               "accessible with Wi-Fi and on-site parking.",
    "std-157": "By far the most-reviewed listing in our entire directory: "
               "6,696 Google reviews at 4.8★, with an extraordinary 25,900 "
               "photos documenting the space. Its Google listing also covers "
               "a record store, museum and souvenir store alongside the "
               "working studio -- a different kind of destination from a "
               "typical commercial recording space, worth knowing before "
               "booking a session there.",
    "std-169": "548 Google reviews at 4.9★ -- one of the highest review "
               "counts in our directory -- open 24 hours, wheelchair "
               "accessible with Wi-Fi and free parking.",
    "std-170": "4.9★ from 176 Google reviews, open 24 hours. Its listing "
               "also covers music education alongside recording and mixing.",
    "std-171": "4.9★ from 216 Google reviews in Miami, open 24 hours, "
               "wheelchair accessible with Wi-Fi.",
    "std-174": "4.8★ from 159 Google reviews, open 24 hours, with "
               "wheelchair access, Wi-Fi and free parking.",
    "std-176": "A perfect 5★ across 254 Google reviews in Nashville, open "
               "24 hours with Wi-Fi on site.",
    "std-179": "A perfect 5★ across 349 Google reviews -- among the "
               "strongest in Nashville -- and notably tagged as a tourist "
               "attraction on Google alongside its recording work, with 299 "
               "photos, wheelchair access, Wi-Fi and free parking.",
    "std-183": "4.9★ from 141 Google reviews with 351 photos -- one of "
               "the most photo-documented studios in our directory. Open 24 "
               "hours, wheelchair accessible with Wi-Fi and on-site parking.",
    "std-184": "4.9★ from 149 Google reviews, with 214 photos on its "
               "listing, wheelchair accessible with Wi-Fi and on-site parking.",
    "std-185": "4.9★ from 120 Google reviews. Beyond recording, its "
               "Google listing also covers coworking space, event venue and "
               "live music venue -- a genuinely multi-purpose space, open 24 "
               "hours with wheelchair access, Wi-Fi and free parking.",
    "std-195": "A perfect 5★ across 101 Google reviews in New Orleans. "
               "Its listing also covers event hosting and music management "
               "alongside recording, with wheelchair access, Wi-Fi and free "
               "parking.",
    "std-204": "4.9★ from 106 Google reviews in New York, open 24 hours, "
               "wheelchair accessible with Wi-Fi.",
    "std-205": "4.9★ from 137 Google reviews, open 24 hours. Also lists "
               "event hosting alongside recording, wheelchair accessible "
               "with Wi-Fi.",
    "std-207": "415 Google reviews at 4.9★ -- a strong review count for "
               "New York -- open daily until 3 AM. Its listing also covers "
               "photography alongside recording, wheelchair accessible with "
               "Wi-Fi.",
    "std-220": "823 Google reviews at 4.9★ and 1,400 photos on its "
               "listing -- one of the most-reviewed and most "
               "photo-documented studios in our entire directory. Open 24 "
               "hours, wheelchair accessible, with Wi-Fi and free parking.",
    "std-226": "A perfect 5★ across 113 Google reviews. Its listing also "
               "spans video editing and production alongside recording, with "
               "Wi-Fi and free parking.",
    "std-228": "4.9★ from 146 Google reviews, open 24 hours, with Wi-Fi "
               "on site.",
    "std-230": "4.9★ from 146 Google reviews, open daily until 1 AM. Its "
               "listing also covers event hosting and live music alongside "
               "recording.",
    "std-234": "4.8★ from 112 Google reviews, with Wi-Fi on site.",
    "std-242": "4.9★ from 181 Google reviews in Phoenix, open daily 10 AM "
               "to 10 PM. Its listing also covers music management and video "
               "editing alongside recording, wheelchair accessible with Wi-Fi.",
    "std-271": "A perfect 5★ across 149 Google reviews in Seattle, with "
               "Wi-Fi and on-site parking.",
    "std-277": "4.9★ from 173 Google reviews, open 24 hours, wheelchair "
               "accessible with Wi-Fi and on-site parking.",
    "std-278": "4.9★ from 103 Google reviews, open daily 9 AM to 11 PM. "
               "Its listing also covers guitar instruction and video "
               "production alongside recording, wheelchair accessible with "
               "Wi-Fi.",
    "std-279": "4.9★ from 127 Google reviews, with Wi-Fi on site.",
    "std-284": "4.8★ from 152 Google reviews, with 180 photos on its "
               "listing, wheelchair accessible, Wi-Fi and free parking.",
    "std-288": "4.8★ from 100 Google reviews, open 24 hours. Its listing "
               "also covers video production alongside recording, "
               "wheelchair accessible with Wi-Fi.",
    "std-290": "A perfect 5★ across 107 Google reviews in Washington, "
               "D.C. Its listing also covers film and video production "
               "alongside music recording, with Wi-Fi on site.",
}


# ---------------------------------------------------------------------------
# STUDIOS  (Google Places crawler)
# ---------------------------------------------------------------------------

def build_studios(base_dir):
    raw_places = load_json_files(FOLDERS["studios"], base_dir)
    seen_place_ids = set()
    kept = []

    for place in raw_places:
        if place.get("permanentlyClosed") or place.get("temporarilyClosed"):
            continue
        if place.get("categoryName") != "Recording studio":
            continue

        place_id = place.get("placeId") or place.get("title")
        if place_id in seen_place_ids:
            continue
        seen_place_ids.add(place_id)
        kept.append(place)

    kept.sort(key=lambda p: (p.get("city") or "", -(p.get("totalScore") or 0)))

    entries = []
    for place in kept:
        rating = place.get("totalScore")
        reviews = place.get("reviewsCount") or 0
        city = place.get("city") or ""
        state = place.get("state") or ""

        extra_stats = []
        hours_summary = summarize_opening_hours(place.get("openingHours"))
        if hours_summary:
            extra_stats.append(hours_summary)
        images_count = place.get("imagesCount")
        if images_count:
            extra_stats.append(f"{format_count(images_count)} photos on Google")

        badges = extract_amenity_badges(place.get("additionalInfo"))

        entries.append({
            "id": f"std-{len(entries) + 1:03d}",
            "category": "studios",
            "categoryLabel": "Recording Studios & Spaces",
            "title": title_case_words(place.get("title")),
            "provider_name": "",
            "location": ", ".join(filter(None, [city, state])),
            "address": place.get("address") or "",
            "phone": place.get("phone") or "",
            "price": "Contact for rates",
            "description": (place.get("description") or "").strip() or
                            f"Recording studio in {city or 'the US'}" + (f", rated {rating}★ from {reviews} reviews." if rating else "."),
            "tags": [t.lower() for t in (place.get("categories") or [])][:4] or ["recording studio"],
            "link": place.get("website") or f"https://www.google.com/maps/place/?q=place_id:{place.get('placeId','')}",
            "linkLabel": "Visit Website" if place.get("website") else "View on Google Maps",
            "image": place.get("imageUrl") or "",
            "rating": rating,
            "reviewsCount": reviews,
            "extraStats": extra_stats,
            "badges": badges,
            "featured": (rating or 0) >= 4.8 and reviews >= 100,
            "sample": False,
        })

    return entries


# ---------------------------------------------------------------------------
# VERIFIED MANUAL FIVERR AFFILIATE LINKS
# ---------------------------------------------------------------------------

# Verified Fiverr affiliate deep-links, generated manually through the Fiverr
# affiliate dashboard's own 'LP URL' tool (Marketing Tools -> Default and Deep
# Links) by the site owner on 2026-08-18. These are known-correct/working links
# (unlike the auto-reconstructed ones from make_fiverr_affiliate_link(), which
# Fiverr does not always honor). Applied as overrides after building engineers/
# visuals so future re-runs of this script do not lose them.
# Listings manually removed after verification found the underlying Fiverr
# gig no longer exists (seller/gig deleted). Filtered out of the final output
# in main(), after all categories are built, so it never shifts the sequential
# IDs of any other listing.
REMOVED_LISTING_IDS = {
    "vis-074",  # user05663302 / "design any cyberpunk art cover" - gig no longer exists on Fiverr (confirmed 2026-08-18)
}

MANUAL_FIVERR_LINK_OVERRIDES = {
    # --- round 2: corrected 2026-08-18 after typo/non-affiliate-link fixes ---
    "eng-006": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Felemproducer%252Fmix-and-master-your-song-to-the-industry-standard",
    "eng-011": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fthedripprovider%252Fprofessionally-mix-and-master-a-rap-or-hip-hop-track-for-you",
    "eng-015": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fenzosilvero%252Fmix-and-master-songs-to-make-them-sound-amazing",
    "eng-018": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fjacksong1022%252Ftune-up-your-vocals",
    "vis-024": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fitommyfrank%252Fcreate-a-cyberpunk-or-sci-fi-illustration-for-you",
    "vis-031": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Faimasterdesign%252Fdesign-high-quality-historical-romance-or-fantasy-cover-fast",
    "vis-079": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fiamaerego%252Fdesign-professional-rap-and-hip-hop-album-cover-art",
    "eng-001": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fneksofficial%252Fprofessionally-produce-and-master-song-or-clean-track-of-unwanted-noise-in-24h",
    "eng-002": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Ferafoos%252Fmix-and-master-your-song-in-3-days",
    "eng-003": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fglastudios%252Fprofessionally-mix-and-master-rap-vocals",
    "eng-004": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fmadushdiwya%252Fprofessional-audio-mixing-and-mastering-for-your-music",
    "eng-005": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fkagenmusic%252Fprofessionally-mix-and-master-your-song",
    "eng-007": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fcarrotymusic%252Fprofessional-mix-and-mastering-streaming-standard",
    "eng-008": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fbrendanreza%252Fskillfully-mix-your-songs-to-sweet-sweet-perfection",
    "eng-009": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fmixedbyhill%252Fmaster-your-music-at-red-factory-studios",
    "eng-010": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fdankjellberg%252Fdo-pro-mixing-an-mastering-to-your-tracks-and-songs",
    "eng-012": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fdespiczeljko%252Fmaster-your-song-pop-rock-trap-edm-any-genre",
    "eng-013": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fjoshermes%252Fmix-and-master-your-hip-hop-trap-rap-song",
    "eng-014": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fbementallyrich%252Fdo-mix-and-master-to-your-song-in-my-professional-studio",
    "eng-016": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Falfredoreed%252Fmix-and-mastering-your-song",
    "eng-017": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Falexreverberi%252Fmix-and-master-your-urban-hiphop-rap-or-solo-singer-song",
    "eng-019": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fsilasbeats%252Fdo-same-day-mastering",
    "eng-020": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fmartinetestudio%252Fdo-a-pro-analog-mastering-to-your-mix",
    "eng-021": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Ftherealbeanzy%252Fbe-your-hip-hop-mixing-engineer",
    "eng-022": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fprodbyrufuh%252Fmix-and-master-your-chill-drill-trap-uk-drill-song",
    "eng-023": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fpinavanucci%252Fmake-a-techno-song-for-you",
    "eng-024": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fdrtinkler%252Fmix-and-master-your-rap-or-hip-hop-track",
    "eng-025": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fjaitay%252Fbe-mixing-and-mastering-your-rap-song-with-pitch-correction",
    "eng-026": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fsoufiandesigner%252Fmake-rap-trap-beat-in-less-than-12-hours",
    "eng-027": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fgabrieleim%252Fmix-and-master-your-song-in-a-clean-progessional-way",
    "eng-028": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fdoubledstudio%252Fmix-and-master-your-modern-trap-urban-song",
    "eng-029": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fzaidiboybeats%252Fmake-fire-melody-for-you",
    "eng-030": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fwiktorcreator%252Fsend-you-more-than-350-professional-fl-studio-vocal-presets",
    "eng-031": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fiamkaiethan%252Fbe-you-hip-hop-vocal-mix-engeneer",
    "eng-032": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fseb_del_vecchio%252Fmaster-your-song-ready-for-digital-stores",
    "eng-033": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Ftonizmusic%252Fmix-your-reggaeton-or-latin-trap-vocals-professionally",
    "eng-034": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Falexrivermusics%252Fbe-your-male-rapper-producer-and-mixing-engineer-for-rap-or-trap",
    "eng-035": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fjoeprod17%252Fprofessionally-mix-your-rap-trap-drill-or-reggaeton-vocals",
    "eng-036": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fvicentedelsante%252Fmix-your-rap-trap-or-drill-vocals-professionally",
    "eng-037": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Faudioaddict596%252Fbe-your-trap-metal-and-phonk-mixing-and-mastering-engineer",
    "vis-001": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fmysellertop%252Fdesign-amazing-flyers-brochures",
    "vis-002": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fsummydesigns%252Fdesign-a-mixtape-or-album-cover",
    "vis-003": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Ftoinggraphicss%252Fdo-advanced-quality-mixtape-covers-and-album-covers",
    "vis-004": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fmirkodellamonic%252Fdesign-album-or-single-cover-art",
    "vis-005": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fflorinbrl%252Fdesign-a-cover-for-your-music-release",
    "vis-006": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Frichardesigns%252Fdesign-your-album-cover-ce4b",
    "vis-007": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fspncrrbns%252Fcraft-your-individual-music-artwork",
    "vis-008": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fkmdesignz%252Fdesign-a-stunning-album-art-or-cd-artwork",
    "vis-009": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Flavidhiarsyad%252Fcollage-artwork-for-albums-cover-etc",
    "vis-010": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fsrdesign4u%252Fdesign-cd-covers-mixtapes-or-flyers",
    "vis-011": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fcrystalgfx%252Fdesign-awesome-single-or-mixtape-cover",
    "vis-012": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Feemran910%252Fdesign-wonderful-album-art-or-cover",
    "vis-013": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fdamagegliters%252Fdo-awesome-mixtape-covers-or-album-covers",
    "vis-014": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fillevenx%252Fcreate-amazing-cover-art-for-your-album-mixtape-ep-song",
    "vis-015": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Femoca_design%252Fdesign-your-music-album-cover-art-or-music-song-artwork",
    "vis-016": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fwasan_seoexpert%252Fdesign-attractive-mixtape-cover-and-album-cover-art",
    "vis-017": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fchapeldesigns%252Fdesign-experimental-abstract-cover-art-for-your-single-or-album",
    "vis-018": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Foleksandrkurta%252Fdesign-album-cover-or-single-cover-or-music-artwork",
    "vis-019": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fsujith_s%252Fdesign-mixtape-covers-album-covers-and-single-covers",
    "vis-020": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fviduboy%252Fdesign-album-or-single-cover-art-28e1",
    "vis-021": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fkgm1988%252Fcreate-digital-collage-album-cover-art-and-poster",
    "vis-022": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fnipz2003%252Fmixtape-cover-art-album-cover-art-or-single-cover",
    "vis-023": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fdegeha%252Fcreate-trippy-psychedelic-surreal-illustration-album-band",
    "vis-025": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fzalmiart%252Fdesign-fantasy-sci-fi-horror-and-dystopian-book-cover-art",
    "vis-026": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Flogo_champion%252Fdo-a-mixtape-cover-design-album-cover-or-flyer-c74a",
    "vis-027": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fhumanoire%252Fcreate-a-high-quality-music-video",
    "vis-028": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fstndgraphics%252Fprepare-layouts-for-music-vinyl-records",
    "vis-029": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fjiyaakhtar%252Fdesign-music-album-cover-single-or-mixtape-artwork",
    "vis-030": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fendryjay%252Fcreate-vaporwave-aesthetic-lofi-vintage-artwork",
    "vis-032": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fdilanpramoda94%252Fdesign-nft-art-or-cover-art",
    "vis-033": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fkxrvm07%252Fdraw-single-cover-or-mixtape-cover",
    "vis-034": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fgloomink%252Fcreate-a-custom-awesome-illustration-for-t-shirt-design",
    "vis-035": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fcloudbear59%252Fdo-design-a-hip-pop-single-cover-or-album-cover",
    "vis-036": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fxgoist_%252Fdesign-unique-album-cover-art-for-your-music-release",
    "vis-037": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fshumaila1408%252Fdesign-unique-amazing-and-brilliant-album-cover-music-or-mixtape-cover",
    "vis-038": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Ffarwa_desin_art%252Fdesign-unique-cd-podcast-mixtape-single-album-cover",
    "vis-039": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Ftazp27%252Fdesign-awesome-custom-artwork-for-your-album-cover",
    "vis-040": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fsabriabdo%252Fcreate-your-portrait-as-a-cartoon",
    "vis-041": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fjasonmgraphics%252Fstunning-album-cover-or-ep-single-cover-art",
    "vis-042": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fbonaedit%252Fdesign-a-stunning-album-or-single-cover-art",
    "vis-043": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fcolorscape%252Fdesign-concept-art-and-photo-real-environments",
    "vis-044": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fgraphicspro255%252Fdraw-a-cartoon-album-cover-art-for-your-music",
    "vis-045": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fbijoysana12%252Fdesign-a-professional-metal-album-cover-artwork-f322",
    "vis-046": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fbrookebagga%252Fcreate-a-custom-3d-vj-loop",
    "vis-047": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fnibera%252Fdesign-a-professional-album-cover-for-your-music",
    "vis-048": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fnicojude%252Fcreate-cinematic-ai-music-videos-animated-visualizers-and-3d-lyric-videos",
    "vis-049": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Flizbecoetzee%252Fillustrate-a-minimal-custom-book-cover-for-any-genre",
    "vis-050": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fpremiumcoverlab%252Fdesign-unique-album-or-single-cover-art",
    "vis-051": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fwalilee%252Fdesign-you-rap-album-cover-or-mixtape-cover-or-ep-cover",
    "vis-052": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fstudioberiman%252Fmake-amazing-custom-illustration-for-your-brand-band-etc",
    "vis-053": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Froteker%252Fcreate-dark-gothic-post-punk-industrial-music-video-for-doomer-obscure-song",
    "vis-054": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fterz_gr%252Fdesign-your-spotify-album-cover-art-professionally",
    "vis-055": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fcncreations%252Fcreate-a-dark-metal-hardcore-music-video-for-a-rock-song",
    "vis-056": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fabirhasan911%252Fdo-science-fiction-book-cover-design",
    "vis-057": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fkianusteve%252Fdo-anime-and-cartoon-album-cover-art",
    "vis-058": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fhakusaladin%252Fcreate-ai-generated-metal-music-video",
    "vis-059": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fgalihpanji%252Fcreate-chaotic-dystopian-music-video-for-ebm-metal-post-punk",
    "vis-060": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fdime51%252Fmake-cyberpunk-anime-style-illustration-fo-your",
    "vis-061": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fjassybaby1%252Fdesign-an-awesome-cd-single-music-album-cover",
    "vis-062": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fsehrish1111%252Fscience-fiction-book-cover",
    "vis-063": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Ftudorslash%252Fcreate-your-hip-hop-album-cover-art-music-mixtape-artwork",
    "vis-064": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fhoodinhy%252Fdesign-your-music-album-cover-art-or-single-artwork",
    "vis-065": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fwhitejude44%252Fcreate-fast-paced-dark-metal-lyric-music-videos-with-gothic-rock-ai-visuals",
    "vis-066": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fdesolubscomics%252Fdesign-anime-hip-hop-emo-album-cover-art-single-rap-mixtape-rnb-retro-y2k-ep",
    "vis-067": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fawanyustira%252Fmake-a-dystopian-music-video-for-metal-post-punk",
    "vis-068": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fheryohsams%252Fdesign-anime-cartoon-album-cover-art-in-hiphop-rap-lofi-edm-funk-or-retro-style",
    "vis-069": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fjundlunn%252Fdesign-custom-anime-manga-album-cover-art-in-hip-hop-rap-lofi-emo-retro-style",
    "vis-070": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fmilestha%252Fcreate-dystopian-dark-music-video-with-gothic-lyric-animation-cinematic-visuals",
    "vis-071": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fberrybruce4%252Fl-produce-a-cinematic-music-video-with-dark-artistic-visuals",
    "vis-072": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Ffave543%252Fdesign-steam-capsule-cyberpunk-art-banner-cover-concept-art-game-poster-arpg",
    "vis-073": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Flusteamart%252Fcreate-scifi-game-steam-capsule-art-fps-game-key-art-cyberpunk-cover-art-banner",
    "vis-075": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Faldermanma%252Fmake-cyberpunk-character-art",
    "vis-076": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fcristhianlondon%252Fillustrate-your-kids-book-with-a-magical-unique-touch",
    "vis-077": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fjoanmpbell%252Fdesign-anime-hip-hop-emo-album-cover-art-single-rap-mixtape-rnb-retro-y2k-ep",
    "vis-078": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fspipau%252Fdo-premium-hip-hop-album-cover-design",
    "vis-080": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fnuvisuals%252Fdesign-a-professional-hip-hop-album-cover-art",
    "vis-081": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fhebause%252Fdo-premium-hip-hop-album-cover-design",
    "vis-082": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fpeapilin%252Fdesign-professional-rap-and-hip-hop-album-cover-art",
    "vis-083": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fkadrearaord%252Fdo-professional-hip-hop-album-cover-design",
    "vis-084": "https://go.fiverr.com/visit/?bta=45990&brand=fiverrmarketplace&landingPage=https%253A%252F%252Fwww.fiverr.com%252Fbarbaobert%252Fcreate-hip-hop-rap-and-trap-album-cover-design",
}


# ---------------------------------------------------------------------------
# EDITORIAL CURATOR NOTES (producers pilot)
# ---------------------------------------------------------------------------

# Short, honest "why this listing is here" notes for each of the 21 curated
# producer profiles. Every sentence is grounded strictly in verifiable data
# already present in the listing itself (follower count, play count, likes,
# upload date, location, tags, description) -- nothing is invented. Profiles
# with very weak or zero real engagement (e.g. prod-006) get a plainly
# neutral, factual note instead of manufactured enthusiasm, per an explicit
# decision to keep every producer in the directory while staying honest about
# thin traction. Applied after producers are built in main() so it never
# affects sequential ID assignment.
PRODUCER_CURATOR_NOTES = {
    "prod-001": "An Atlanta-based beatmaker with a catalog dating back to 2016. "
                "Engagement on this particular upload is modest, so we'd suggest "
                "sampling the full SoundCloud profile before reaching out.",
    "prod-002": "An Atlanta trap producer whose catalog has crossed 1,000 plays "
                "and built a small, steady following since 2017.",
    "prod-003": "Based in the Bankhead area of Atlanta with a following of over "
                "370 on SoundCloud, though this specific 2013 upload has seen "
                "limited plays -- worth checking the artist's newer material too.",
    "prod-004": "An Atlanta producer whose 2014 upload has passed 3,800 plays and "
                "56 likes -- solid organic traction for a single track.",
    "prod-005": "Atlanta-based, with 470 SoundCloud followers. This beat leans "
                "into a Russian-rap-influenced style (referencing Скриптонит and "
                "Truwer), showing range beyond standard US trap.",
    "prod-006": "A newly uploaded catalog (September 2025) with no engagement "
                "data yet. There isn't a track record to point to, but it's kept "
                "in the directory for those exploring newer Atlanta trap producers.",
    "prod-007": "An Atlanta-based beatmaker with a modest SoundCloud footprint "
                "(58 followers). Engagement on this catalog is still limited.",
    "prod-008": "A hip-hop remix specialist -- this rework of Mobb Deep's 'Shook "
                "Ones Pt. 2' has drawn 538 plays, showing an ear for classic "
                "boom-bap source material.",
    "prod-009": "A boom-bap producer with a following approaching 900 on "
                "SoundCloud and steady engagement (41 likes) on this instrumental.",
    "prod-010": "The standout profile in this category: a free-use boom-bap "
                "instrumental out of Medellín, Colombia that has racked up over "
                "80,000 plays and 1,448 likes -- real, verifiable traction well "
                "ahead of every other producer listed here.",
    "prod-011": "371 SoundCloud followers, though engagement on this specific "
                "upload is light (2 likes, 258 plays) -- worth sampling the track "
                "directly before deciding.",
    "prod-012": "A Morocco-based dark-trap producer who publishes exact BPM and "
                "key details with each beat (this one: 132 BPM, key of B), which "
                "makes tempo-matching straightforward for artists.",
    "prod-013": "Canada-based, with a Mac Miller x Wiz Khalifa-style piano beat "
                "that has drawn over 3,400 plays and 62 likes -- one of the "
                "stronger engagement numbers in this category.",
    "prod-014": "A Belgium-based trap producer sourced directly from SoundCloud. "
                "A smaller catalog so far, worth an audition before booking.",
    "prod-015": "A Nigeria-based producer offering free-use trap instrumentals in "
                "a Gunna/Young Thug-adjacent style -- a reasonable starting point "
                "for testing a sound before commissioning custom work.",
    "prod-016": "Based in Bahia, Brazil, with a Travis Scott x Orochi-styled trap "
                "beat that has passed 6,100 plays and 113 likes -- one of the "
                "more engaged profiles in this category.",
    "prod-017": "Accra, Ghana-based, blending Afrobeat influences into hip-hop "
                "instrumentals (this beat is styled after Sarkodie) -- a distinct "
                "regional sound within the directory.",
    "prod-018": "A trap/instrumental producer specializing in darker-toned "
                "beats. A smaller catalog so far, best judged by listening "
                "directly to the track.",
    "prod-019": "Based in Rio de Janeiro, Brazil, with 1,200+ SoundCloud "
                "followers and 3,300 plays on this YG-style trap beat -- a "
                "well-established international profile.",
    "prod-020": "A free-use, old-school boom-bap instrumental that has picked up "
                "over 5,500 plays and 76 likes -- strong organic traction for a "
                "freely licensable beat.",
    "prod-021": "An old-school boom-bap producer with a modest but genuine "
                "following. A smaller catalog, worth sampling directly.",
}


# ---------------------------------------------------------------------------
# FAQ GENERATION (per-listing, grounded strictly in real fields)
# ---------------------------------------------------------------------------
#
# Every question below is only included when there is real data on the entry
# to answer it honestly -- a listing with thinner data simply gets fewer FAQ
# entries rather than a fabricated or generic-sounding answer. This mirrors
# the same honesty rule used for curatorNote. Answers target the exact
# long-tail questions a real visitor would type into Google ("how much does
# it cost to...", "is X a verified seller", "what are the opening hours of
# X"), which is also what FAQPage structured data is designed to surface in
# search results.

def _faq_producer(entry):
    provider = entry.get("provider_name") or "this producer"
    faq = []

    faq.append({
        "question": f"How much does it cost to license a beat from {provider}?",
        "answer": (f"{provider} lists pricing as \"Contact for licensing\" rather than a fixed "
                   f"rate -- reach out directly via their SoundCloud profile (linked above) to "
                   f"discuss usage rights and pricing."),
    })

    tags = entry.get("tags") or []
    if tags:
        faq.append({
            "question": "What style or genre is this beat?",
            "answer": f"This track is tagged {', '.join(tags)} on SoundCloud.",
        })

    extra_stats = entry.get("extraStats") or []
    likes = entry.get("likes")
    if extra_stats or likes is not None:
        stats_bits = [f"{likes} like{'s' if likes != 1 else ''}"] if likes is not None else []
        stats_bits += extra_stats
        faq.append({
            "question": f"How much real engagement does this track from {provider} have?",
            "answer": f"As listed on SoundCloud: {', '.join(stats_bits)}.",
        })

    if entry.get("location"):
        faq.append({
            "question": f"Where is {provider} based?",
            "answer": f"{provider} lists their location as {entry['location']} on SoundCloud.",
        })

    if "✓ Verified SoundCloud Account" in (entry.get("badges") or []):
        faq.append({
            "question": f"Is {provider}'s SoundCloud account verified?",
            "answer": f"Yes -- {provider} carries a verified-account badge on SoundCloud.",
        })

    return faq


def _faq_fiverr(entry):
    noun = "engineer" if entry.get("category") == "engineers" else "designer"
    provider = entry.get("provider_name") or f"this {noun}"
    faq = []

    if entry.get("price"):
        faq.append({
            "question": f"How much does it cost to work with {provider}?",
            "answer": f"Pricing for this gig starts {entry['price']} on Fiverr.",
        })

    extra_stats = entry.get("extraStats") or []
    delivery = next((s for s in extra_stats if "delivery" in s), None)
    if delivery:
        faq.append({
            "question": "How fast is delivery?",
            "answer": f"{provider} lists a {delivery} on this gig.",
        })

    rating = entry.get("rating")
    reviews = entry.get("reviewsCount") or 0
    is_pro = "PRO Verified Seller" in (entry.get("badges") or [])
    if rating and reviews > 0:
        pro_bit = ", and is a Fiverr Pro-verified seller" if is_pro else ""
        faq.append({
            "question": f"Is {provider} a trusted, verified seller?",
            "answer": (f"{provider} holds a {rating}★ rating across {reviews} Fiverr "
                       f"review{'s' if reviews != 1 else ''}{pro_bit}."),
        })
    else:
        faq.append({
            "question": f"Is {provider} a trusted, verified seller?",
            "answer": ("This gig doesn't have Fiverr reviews on record yet -- worth checking "
                       "their full Fiverr profile directly before booking."),
        })

    languages = next((s for s in extra_stats if s.startswith("speaks")), None)
    if languages:
        faq.append({
            "question": f"What languages does {provider} speak?",
            "answer": f"{provider} {languages}, per their Fiverr profile.",
        })

    tags = entry.get("tags") or []
    if tags:
        faq.append({
            "question": f"What genres or styles does {provider} work in?",
            "answer": f"Listed specialties for this gig include: {', '.join(tags)}.",
        })

    return faq


def _faq_studio(entry):
    title = entry.get("title") or "this studio"
    faq = []

    if entry.get("price"):
        faq.append({
            "question": f"How much does it cost to book {title}?",
            "answer": (f"{title} lists pricing as \"Contact for rates\" -- contact them "
                       f"directly (details above) for a quote."),
        })

    extra_stats = entry.get("extraStats") or []
    hours = next((s for s in extra_stats
                  if "AM" in s or "PM" in s or "24 hours" in s or "vary" in s.lower()), None)
    if hours:
        faq.append({
            "question": f"What are {title}'s opening hours?",
            "answer": f"{title}'s listed hours on Google: {hours}.",
        })

    badges = entry.get("badges") or []
    if badges:
        faq.append({
            "question": f"What amenities are available at {title}?",
            "answer": f"Amenities listed on Google for {title}: {', '.join(badges)}.",
        })

    rating = entry.get("rating")
    reviews = entry.get("reviewsCount") or 0
    if rating and reviews > 0:
        faq.append({
            "question": f"Is {title} a highly-rated studio?",
            "answer": (f"{title} holds a {rating}★ rating from {reviews} Google "
                       f"review{'s' if reviews != 1 else ''}."),
        })
    else:
        faq.append({
            "question": f"Is {title} a highly-rated studio?",
            "answer": f"{title} doesn't have Google reviews on record yet.",
        })

    tags = entry.get("tags") or []
    extra_tags = [t for t in tags if t != "recording studio"]
    if extra_tags:
        faq.append({
            "question": f"Does {title} offer anything beyond standard recording sessions?",
            "answer": f"Yes -- its Google listing also covers: {', '.join(extra_tags)}.",
        })

    return faq


def generate_faq(entry):
    """Returns a list of {"question", "answer"} pairs for a listing, built
    entirely from real fields already on the entry -- see the category
    helpers above. Dispatches by category."""
    category = entry.get("category")
    if category == "producers":
        return _faq_producer(entry)
    if category in ("engineers", "visuals"):
        return _faq_fiverr(entry)
    if category == "studios":
        return _faq_studio(entry)
    return []


def load_manual_faq_additions(base_dir):
    """Reads faq_additions.xlsx (id, title, question, answer columns; one row
    per Q&A pair, multiple rows per id allowed) if it exists, and returns a
    dict of id -> [{"question", "answer"}, ...]. This is the durable,
    editable place to add hand-written FAQ paragraphs -- for existing
    listings or ones not yet built -- so that content is never lost or
    redone on a future re-run of this script. Returns {} if the file is
    absent (this feature is entirely optional)."""
    path = os.path.join(base_dir, "faq_additions.xlsx")
    if not os.path.exists(path):
        return {}

    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    sheet = wb["FAQ Additions"] if "FAQ Additions" in wb.sheetnames else wb.active

    headers = [str(c.value).strip().lower() if c.value else "" for c in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        id_col = headers.index("id")
        q_col = headers.index("question")
        a_col = headers.index("answer")
    except ValueError:
        return {}

    additions = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or row[id_col] is None:
            continue
        listing_id = str(row[id_col]).strip()
        question = (row[q_col] or "").strip() if row[q_col] else ""
        answer = (row[a_col] or "").strip() if row[a_col] else ""
        if not listing_id or not question or not answer:
            continue
        additions.setdefault(listing_id, []).append({"question": question, "answer": answer})

    return additions


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    base_dir = os.path.dirname(os.path.abspath(__file__))

    print("Building producers...")
    producers = build_producers(base_dir)
    manual_producers = build_manual_producer_additions(len(producers) + 1)
    producers = producers + manual_producers
    print(f"  -> {len(producers)} curated producer profiles "
          f"({len(manual_producers)} manually sourced via Google discovery)")

    applied_notes = 0
    for entry in producers:
        note = PRODUCER_CURATOR_NOTES.get(entry.get("id"))
        if note:
            entry["curatorNote"] = note
            applied_notes += 1
    print(f"  -> applied {applied_notes} editorial curator notes to producer profiles")

    print("Building sound engineers...")
    engineers = build_fiverr_category(
        base_dir, "engineers", "engineers", "Mixing & Mastering Engineers", "eng",
        require_hip_hop_genre=True, require_engineer_keywords=True,
    )
    print(f"  -> {len(engineers)} curated engineer listings")

    print("Building visual / cover art designers...")
    visuals = build_fiverr_category(
        base_dir, "visuals", "visuals", "Cover Art & Visual Designers", "vis",
        require_hip_hop_genre=False, require_engineer_keywords=False,
        genre_field="musical_genre",
    )
    print(f"  -> {len(visuals)} curated designer listings")

    applied_overrides = 0
    for entry in engineers + visuals:
        override = MANUAL_FIVERR_LINK_OVERRIDES.get(entry.get("id"))
        if override:
            entry["link"] = override
            applied_overrides += 1
    print(f"  -> applied {applied_overrides} verified manual Fiverr affiliate links")

    algo_notes = 0
    featured_notes = 0
    for entry in engineers + visuals:
        entry["curatorNote"] = generate_fiverr_curator_note(entry)
        algo_notes += 1
        manual_note = FIVERR_FEATURED_CURATOR_NOTES.get(entry.get("id"))
        if manual_note:
            entry["curatorNote"] = manual_note
            featured_notes += 1
    print(f"  -> applied {algo_notes} algorithmic curator notes to engineer/designer "
          f"listings ({featured_notes} upgraded to hand-written editorial notes "
          f"for featured listings)")

    print("Building recording studios...")
    studios = build_studios(base_dir)
    print(f"  -> {len(studios)} curated studio listings")

    algo_studio_notes = 0
    featured_studio_notes = 0
    for entry in studios:
        entry["curatorNote"] = generate_studio_curator_note(entry)
        algo_studio_notes += 1
        manual_note = STUDIO_FEATURED_CURATOR_NOTES.get(entry.get("id"))
        if manual_note:
            entry["curatorNote"] = manual_note
            featured_studio_notes += 1
    print(f"  -> applied {algo_studio_notes} algorithmic curator notes to studio "
          f"listings ({featured_studio_notes} upgraded to hand-written editorial "
          f"notes for featured studios)")

    master_data = producers + engineers + visuals + studios

    faq_count = 0
    for entry in master_data:
        entry["faq"] = generate_faq(entry)
        faq_count += len(entry["faq"])
    print(f"  -> generated {faq_count} algorithmic FAQ entries across "
          f"{len(master_data)} listings")

    manual_faq = load_manual_faq_additions(base_dir)
    if manual_faq:
        applied_manual_faq = 0
        for entry in master_data:
            extra = manual_faq.get(entry.get("id"))
            if extra:
                entry["faq"] = entry["faq"] + extra
                applied_manual_faq += len(extra)
        print(f"  -> merged {applied_manual_faq} hand-written FAQ entries from "
              f"faq_additions.xlsx ({len(manual_faq)} listings)")
    else:
        print("  -> no faq_additions.xlsx found (optional -- skipped)")

    before_removal = len(master_data)
    master_data = [x for x in master_data if x.get("id") not in REMOVED_LISTING_IDS]
    if before_removal != len(master_data):
        print(f"  -> removed {before_removal - len(master_data)} discontinued listing(s): "
              f"{sorted(REMOVED_LISTING_IDS)}")

    output_path = os.path.join(base_dir, "directory.json")
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(master_data, f, ensure_ascii=False, indent=2)

    print()
    print(f"Done. Wrote {len(master_data)} total listings to {output_path}")


if __name__ == "__main__":
    main()
